from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http.response import JsonResponse
from core.models import Usuario
from .models import Producto, Categoria
import csv
from django.core.files.storage import FileSystemStorage
from .functions import *
from django.contrib import messages
import openpyxl
from io import BytesIO
from django.http import HttpResponse
# Create your views here.
app_name = 'productos'

# Muestra la tabla de productos.


@login_required
def productos(request):
    autorizado = request.session.get('autorizado')
    usuario_nombre = request.session.get('usuario_nombre')
    img = request.session.get('img')
    prods = Producto.objects.filter(resultado=0).all()
    if request.user.is_superuser or request.user.groups.filter(name='Gerencia').exists():
        autorizado = True
    data = {
        'usuario': usuario_nombre,
        'img': img,
        'prods': prods,
        'autorizado': autorizado,
    }
    return render(request, 'productos/productos.html', data)

# Vista que se usa a traves de una petición AJAX desde el frontend para obtener la información de los usuarios y de los proyectos en formato JSON.
# y utilizarla para completar un modal de edición de productos.


@login_required
def info_editar_producto(request):
    Cate = Categoria.objects.all()

    Cate_data = list(Cate.values('id', 'nombre'))
    data = {
        'Cate': Cate_data,
    }
    return JsonResponse(data)


def obtener_producto(request, producto_id):
    try:
        producto = Producto.objects.get(pk=producto_id)
        data = {
            'nombre': producto.nombre,
            'codigo': producto.codigo,
            'precio': producto.precio,
            'categoria_id': producto.categoria.id,
            'ancho': producto.ancho,
            'alto': producto.alto,
            'factor': producto.factor,
        }
        return JsonResponse(data)
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado.'}, status=404)

# Vista que recibe la información de los modales de editar y agregar producto para luego actualizar la base de datos.


@login_required
def guardar_producto(request):
    if request.method == "POST":
        codigo = request.POST.get('codigo', None)
        nombre = request.POST['nombre']
        precio = request.POST['precio']
        ancho = request.POST['ancho']
        alto = request.POST['alto']
        factor = request.POST.get('factor_edit', 1)
        nueva_categoria = request.POST.get('nueva_categoria', '').strip()
        categoria_id = request.POST.get('categ')
        vendedor = request.session.get('vendedor')
        # Determinar la categoría
        if nueva_categoria:
            # Verificar si la categoría ya existe
            categoria, created = Categoria.objects.get_or_create(
                nombre=nueva_categoria)
        else:
            # Si no se ingresó una nueva categoría, usar la seleccionada del select
            categoria = Categoria.objects.get(
                id=categoria_id) if categoria_id else None

        if codigo:  # Editar producto existente
            producto = get_object_or_404(Producto, codigo=codigo)
            producto.nombre = nombre
            producto.precio = precio
            producto.ancho = ancho
            producto.alto = alto
            producto.factor = factor
            producto.categoria = categoria
            producto.save()
        else:  # Crear nuevo producto
            Producto.objects.create(
                nombre=nombre,
                precio=precio,
                ancho=ancho,
                alto=alto,
                factor=factor,
                categoria=categoria,
                vendedor=Usuario.objects.get(id=vendedor),
            )

        # Redirigir o devolver respuesta
        messages.success(request, "Producto guardado correctamente.")
        return redirect('productos')
    else:
        return JsonResponse({'error': 'Método no permitido'}, status=405)


# Vista que permite borrar un producto de la base de datos.


@login_required
def borrar_producto(request, producto_id):
    try:
        prod = Producto.objects.filter(pk=producto_id)
        prod.delete()
        messages.success(request, 'El producto se ha borrado correctamente.')
    except Exception as e:
        messages.error(
            request, f'No se ha podido borrar el producto. Error({e})')
    return redirect('/productos', messages)

# Función que carga productos desde un archivo excel.


def cargar_productos(request):
    total = 0
    contador = 0
    try:
        if request.method == 'POST' and request.FILES.get('excel_file'):
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(values_only=True):
                total += 1
                if total == 1:
                    continue

                # Verificar si la fila está vacía
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    break  # Detener el procesamiento si la fila está vacía

                values = []
                for col in range(1, 8):
                    cell = sheet.cell(row=total, column=col)
                    value = cell.value if cell.value is not None else " "
                    values.append(str(value))
                codigo, nombre, ancho, alto, precio, factor, categoria = values
                if not comparar(codigo):
                    try:
                        Producto.objects.create(
                            codigo=codigo,
                            nombre=nombre,
                            precio=precio,
                            factor=factor,
                            ancho=ancho if ancho != ' ' else 1,
                            alto=alto if alto != ' ' else 1,
                            categoria=Categoria.objects.get(
                                nombre=categoria),
                        )
                        contador += 1
                    except Exception as e:
                        messages.error(
                            request, f"Sucedió un error inesperado. Error({e})")
            messages.success(
                request, f'Se han cargado {contador} de {total} productos correctamente.')
    except Exception as e:
        messages.error(
            request, f'No se han podido cargar los productos. Error({e})')
    return redirect('/productos', messages)


# Vista que permite general un excel con los productos de la base de datos.
@login_required
def exportar_productos(request):
    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Backup_Lista de productos'

    # Definir encabezados
    headers = ['Código', 'Nombre', 'Ancho',
               'Alto', 'Precio', 'Factor', 'Categoría']
    sheet.append(headers)

    # Agregar datos de los productos
    productos = Producto.objects.all().filter(presupuesto=None)
    for producto in productos:
        data = [
            producto.codigo,
            producto.nombre,
            producto.ancho,
            producto.alto,
            producto.precio,
            producto.factor,
            producto.categoria.nombre
        ]
        sheet.append(data)

    # Guardar el libro en un buffer de memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Crear la respuesta HTTP
    response = HttpResponse(
        buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'

    return response

# Vista que permite borrar todos los productos.


def borrar_todos_productos(request):
    try:
        Producto.objects.all().filter(presupuesto=None).filter(vendedor=None).delete()
        messages.success(
            request, 'Todos los productos se han borrado correctamente.')
    except Exception as e:
        messages.error(
            request, f'No se han podido borrar todos los productos. Error({e})')
    return redirect('/productos', messages)


@login_required
def categorias(request):
    autorizado = request.session.get('autorizado')
    usuario_nombre = request.session.get('usuario_nombre')
    img = request.session.get('img')
    categorias = Categoria.objects.all()

    data = {
        'usuario': usuario_nombre,
        'autorizado': autorizado,
        'img': img,
        'categorias': categorias,
    }

    return render(request, 'productos/categorias.html', data)


def obtener_categoria(request, categoria_id):
    try:
        categoria = Categoria.objects.get(id=categoria_id)
        return JsonResponse({'id': categoria.id, 'nombre': categoria.nombre})
    except Categoria.DoesNotExist:
        return JsonResponse({'error': 'Categoría no encontrada'}, status=404)


def guardar_categoria(request):
    if request.method == 'POST':
        categoria_id = request.POST.get('id')
        nombre = request.POST.get('nombre')

        if categoria_id:  # Si hay un ID, estamos en modo edición
            try:
                categoria = Categoria.objects.get(id=categoria_id)
                categoria.nombre = nombre
                categoria.save()
                messages.success(
                    request, f'Categoría "{nombre}" actualizada con éxito.')
            except Categoria.DoesNotExist:
                messages.error(request, 'La categoría no existe.')
        else:  # No hay ID, creamos una nueva categoría
            try:
                Categoria.objects.create(nombre=nombre)
                messages.success(
                    request, f'Nueva categoría "{nombre}" creada con éxito.')
            except Exception as e:
                messages.error(
                    request, f'Error al crear la categoría: {str(e)}')

        return redirect('categorias')  # Redirigir a la vista de categorías

    return redirect('categorias')  # Si no es POST, redirigir


def borrar_categoria(request, categoria_id):
    try:
        cat = Categoria.objects.get(id=categoria_id)
        cat.delete()
        messages.success(request, 'La categoria se ha borrado correctamente.')
    except Exception as e:
        messages.error(
            request, f'No se ha podido borrar la categoria. Error({e})')
    return redirect('/productos/categorias', messages)
